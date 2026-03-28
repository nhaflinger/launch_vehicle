#!/usr/bin/env python3
"""
Launch Vehicle Design Spreadsheet Generator
Usage: python3 generate_lv_design.py
Output: launch_vehicle_design.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Constants ─────────────────────────────────────────────────────────────────
OUTPUT_FILE = "launch_vehicle_design.xlsx"
G0 = 9.80665
MAX_STAGES = 3

# ── Colors ────────────────────────────────────────────────────────────────────
C = {
    "hdr":    "1F3864",   # dark blue header
    "sub":    "2F5496",   # medium blue subheader
    "input":  "FFF2CC",   # yellow: user editable
    "calc":   "DAEEF3",   # light blue: formula/calculated
    "sec":    "D6DCE4",   # gray: section label
    "stage":  ["E2EFDA", "FCE4D6", "EAD1DC"],  # per-stage tints
    "pos":    "C6EFCE",
    "neg":    "FFC7CE",
    "white":  "FFFFFF",
    "ltgray": "F2F2F2",
}

# ── Engine Database ───────────────────────────────────────────────────────────
# name, vehicle_application, propellant, isp_vac_s, isp_sl_s,
# thrust_vac_kn, thrust_sl_kn, mass_kg, chamber_press_bar, exp_ratio, notes
ENGINES = [
    # LOX/RP-1
    ("Merlin 1D",       "Falcon 9 S1",          "LOX/RP-1",  311, 282,  934,  845,  470, 97,  16,   "9 engines on F9 first stage; sea-level optimized"),
    ("Merlin 1D Vac",   "Falcon 9 S2",          "LOX/RP-1",  348, None, 934,  None, 490, 97,  165,  "Vacuum-optimized fixed nozzle extension"),
    ("RD-180",          "Atlas V CBC",           "LOX/RP-1",  338, 311, 4152, 3828, 5480,257, 36.4, "Dual-chamber single-turbopump; Russian-designed"),
    ("RD-107A",         "Soyuz Strap-on",        "LOX/RP-1",  320, 263, 1020,  838, 1190, 60, 21.0, "4 strap-on boosters; 4 fixed + 4 vernier chambers"),
    ("RD-108A",         "Soyuz Core",            "LOX/RP-1",  320, 257,  990,  792, 1278, 60, 21.0, "Core stage; 4 main + 4 vernier chambers"),
    ("RD-0124",         "Soyuz S3 / Angara S2",  "LOX/RP-1",  359, 311,  294,  None, 480,157, 30.0, "High-performance staged combustion upper stage"),
    ("F-1",             "Saturn V S-IC",         "LOX/RP-1",  304, 263, 7740, 6770, 8391, 70, 16.0, "5 engines on Saturn V first stage; largest US engine"),
    ("RD-191",          "Angara S1",             "LOX/RP-1",  337, 311, 2090, 1922, 2200,263, 37.0, "Single-chamber RD-180 derivative"),
    ("NK-33 (AJ-26)",   "Antares S1",            "LOX/RP-1",  331, 297, 1638, 1510, 1222,146, 27.0, "Soviet N1 engine; modernized as AJ-26"),
    ("Rutherford",      "Electron S1",           "LOX/RP-1",  311, 303,   24,   22,   35,145, 11.0, "9 engines; electric-pump cycle; additive manufactured"),
    ("Rutherford Vac",  "Electron S2",           "LOX/RP-1",  343, None,  25, None,   35,145, 250,  "Vacuum variant; electric-pump cycle"),
    # LOX/LH2
    ("RS-25 (SSME)",    "SLS Core / Shuttle",    "LOX/LH2",   452, 366, 2090, 1860, 3177,206, 69.0, "3x on SLS; 4x on Shuttle; high-performance SSME"),
    ("RL-10A-4-2",      "Atlas V Centaur",       "LOX/LH2",   451, None,  99, None,  168, 28, 84.0, "Expander cycle; multiple restarts"),
    ("RL-10B-2",        "Delta IV DCSS",         "LOX/LH2",   462, None, 110, None,  277, 44,285,   "Extendable carbon-carbon nozzle"),
    ("RL-10C-1",        "Vulcan Centaur",        "LOX/LH2",   450, None, 102, None,  190, 34, 84.0, "Updated RL-10; simplified nozzle"),
    ("RL-10C-3",        "SLS ICPS",              "LOX/LH2",   460, None, 110, None,  190, 34,150,   "High-expansion ratio variant"),
    ("Vulcain 2",       "Ariane 5 EPC",          "LOX/LH2",   431, 318, 1340,  960, 1686,115, 60.0, "Gas-generator cycle; Ariane 5 core"),
    ("Vinci",           "Ariane 6 ULPM",         "LOX/LH2",   465, None, 180, None,  550, 61,240,   "Expander cycle; restartable upper stage"),
    ("HM7B",            "Ariane 5 ESC-A",        "LOX/LH2",   446, None,  65, None,  165, 35, 83.0, "Gas-generator; multiple restarts"),
    ("J-2",             "Saturn V S-II / S-IVB", "LOX/LH2",   421, None,1033, None, 1788, 52, 27.5, "5x on S-II; 1x on S-IVB; restartable"),
    ("RS-68A",          "Delta IV Heavy CBC",    "LOX/LH2",   412, 365, 3370, 2949, 6597, 97, 21.5, "Gas-generator; simple design; largest H2 engine"),
    ("BE-3U",           "New Glenn S2",          "LOX/LH2",   440, None, 710, None, 2700, 90,200,   "Blue Origin upper stage engine; expander cycle; 2x on NG 7x2"),
    ("BE-3U Block 2",   "New Glenn Block 2 S2",  "LOX/LH2",   440, None, 890, None, 2700, 90,200,   "Upgraded BE-3U; 890 kN vac (200,000 lbf) vs 710 kN original; 4x on NG 9x4; Isp unchanged"),
    # LOX/CH4
    ("Raptor 2",        "Starship / Super Heavy","LOX/CH4",   350, 327, 2350, 2200, 1500,330, 32.0, "Full-flow staged combustion; 33x on Super Heavy"),
    ("Raptor Vac",      "Starship upper",        "LOX/CH4",   380, None,2200, None, 1600,330,107,   "Vacuum variant; large fixed nozzle; 3x on Starship"),
    ("BE-4",            "Vulcan / New Glenn S1", "LOX/CH4",   339, 310, 2400, 2100, 4760,134, 40.0, "Oxygen-rich staged combustion; 2x Vulcan / 7x NG 7x2 baseline"),
    ("BE-4 Block 2",    "New Glenn Block 2 S1",  "LOX/CH4",   339, 310, 2847, 2490, 4760,134, 40.0, "Upgraded BE-4; 2,847 kN SL (640,000 lbf) with propellant subcooling; vac est. proportional; same Isp"),
    # Hypergolic
    ("AJ-10-190",       "Orion SPS / STS OMS",   "NTO/MMH",   316, None,  27, None,  118,  9, 49.0, "Pressure-fed; restartable; long heritage"),
    ("Vikas",           "PSLV / GSLV S2",        "NTO/UDMH",  293, 270,  800,  725,  634, 58, 16.0, "Indian engine; Ariane Viking derivative"),
    ("YF-20 (cluster)", "Long March S1",         "NTO/UDMH",  289, 259, 2962, 2799, 4000, 65, 14.0, "4-chamber cluster used on CZ-2/3/4 first stages"),
    ("Aestus",          "Ariane 5 ATV",          "NTO/MMH",   324, None,  29, None,  111, 11, 84.0, "Pressure-fed; restartable; ESA upper stage"),
    # Solid
    ("SRB (SLS 5-seg)", "SLS Block 1",           "Solid HTPB",269, 242,16000,14900,90718, 62,  7.5, "Largest SRB ever flown; 2x on SLS"),
    ("RSRM (Shuttle)",  "Space Shuttle",         "Solid HTPB",268, 242,14678,12453,90714, 62,  7.7, "4-segment Shuttle SRB"),
    ("GEM-63XL",        "Vulcan strap-on",       "Solid HTPB",279, 252, 2140, 1930, 7000, 60, 11.0, "Solid strap-on; also used on Atlas V 500"),
    ("P80",             "Ariane 6 S1",           "Solid HTPB",278, 255, 3040, 2820, 5650, 95, 16.0, "Ariane 6 first stage solid motor"),
    ("Zefiro-23",       "Vega S2",               "Solid HTPB",287, None,1120, None, 1315, 87, 30.0, "Vega second stage solid motor"),
    ("Zefiro-9",        "Vega S3",               "Solid HTPB",296, None, 314, None,  260, 98, 54.0, "Vega third stage solid motor"),
    ("Star-48BV",       "Various kick stages",   "Solid HTPB",292, None,  68, None,  119, 40, 55.0, "Small solid kick motor; spin-stabilized"),
    ("Castor 30B",      "Antares S2",            "Solid HTPB",302, None, 357, None, 1437, 80, 57.0, "Antares second stage solid motor"),
]

# ── Reference Vehicle Database ─────────────────────────────────────────────────
# Each vehicle has metadata and a list of stages:
#   (stage_name, engine_name, n_engines, propellant_mass_kg, dry_mass_kg, notes)
VEHICLES = [
    {
        "name": "Falcon 9 Block 5", "country": "USA", "operator": "SpaceX",
        "first_flight": 2018, "status": "Active",
        "payload_leo_kg": 22800, "payload_gto_kg": 8300, "payload_tli_kg": None,
        "glow_kg": 549054, "height_m": 70.0, "diameter_m": 3.66,
        "notes": "22,800 kg LEO expendable / 17,400 kg reusable; partially reusable first stage",
        "stages": [
            ("S1 Core",     "Merlin 1D",      9, 411000, 22200, "Reusable; includes grid fins, landing legs, interstage"),
            ("S2 Upper",    "Merlin 1D Vac",  1, 111500,  4000, "Single Merlin Vacuum engine"),
        ],
    },
    {
        "name": "Falcon Heavy", "country": "USA", "operator": "SpaceX",
        "first_flight": 2018, "status": "Active",
        "payload_leo_kg": 63800, "payload_gto_kg": 26700, "payload_tli_kg": None,
        "glow_kg": 1420788, "height_m": 70.0, "diameter_m": 12.2,
        "notes": "Expendable config; 3 F9 cores; center core reinforced",
        "stages": [
            ("Center Core", "Merlin 1D",      9, 411000, 25600, "Reinforced; cross-feed ready"),
            ("Side Booster (x2)", "Merlin 1D",9, 411000, 22200, "Two identical; mass shown per booster"),
            ("S2 Upper",    "Merlin 1D Vac",  1, 111500,  4000, "Single Merlin Vacuum"),
        ],
    },
    {
        "name": "Atlas V 401", "country": "USA", "operator": "ULA",
        "first_flight": 2002, "status": "Active (limited)",
        "payload_leo_kg": 9750, "payload_gto_kg": 4750, "payload_tli_kg": None,
        "glow_kg": 334500, "height_m": 58.3, "diameter_m": 3.81,
        "notes": "401 = 4m fairing, 0 SRBs, 1 RL-10; 551 config adds 5 GEM SRBs and dual RL-10",
        "stages": [
            ("CBC Booster",  "RD-180",        1, 284089, 21054, "Russian-built engine; kerosene/LOX"),
            ("Centaur III",  "RL-10A-4-2",    1,  20830,  2316, "High-energy upper stage; LOX/LH2"),
        ],
    },
    {
        "name": "Delta IV Heavy", "country": "USA", "operator": "ULA",
        "first_flight": 2004, "status": "Retired 2024",
        "payload_leo_kg": 28370, "payload_gto_kg": 14220, "payload_tli_kg": None,
        "glow_kg": 733400, "height_m": 72.0, "diameter_m": 15.0,
        "notes": "3 identical CBCs; heaviest US expendable before SLS",
        "stages": [
            ("Center CBC",   "RS-68A",         1, 202000, 26760, "LOX/LH2 core; mass per CBC"),
            ("Side CBC (x2)","RS-68A",         1, 202000, 26760, "Two identical side boosters; mass per booster"),
            ("DCSS",         "RL-10B-2",       1,  27220,  3490, "Delta Cryogenic Second Stage; LOX/LH2"),
        ],
    },
    {
        "name": "Ariane 5 ECA", "country": "Europe", "operator": "Arianespace",
        "first_flight": 2002, "status": "Retired 2023",
        "payload_leo_kg": 21000, "payload_gto_kg": 10865, "payload_tli_kg": None,
        "glow_kg": 777000, "height_m": 59.0, "diameter_m": 5.4,
        "notes": "2 EAP solid strap-ons + EPC LOX/LH2 core + ESC-A LOX/LH2 upper",
        "stages": [
            ("EAP Solids (x2)","P80",          1, 237000, 30000, "Solid strap-on; each booster; actual Isp ~274s; P80 shown as proxy"),
            ("EPC Core",     "Vulcain 2",       1, 158000, 14700, "LOX/LH2 core stage"),
            ("ESC-A Upper",  "HM7B",           1,  14700,  1400, "LOX/LH2 upper stage; restartable"),
        ],
    },
    {
        "name": "Soyuz-2.1b", "country": "Russia", "operator": "Roscosmos",
        "first_flight": 2006, "status": "Active",
        "payload_leo_kg": 8200, "payload_gto_kg": 2800, "payload_tli_kg": None,
        "glow_kg": 312000, "height_m": 46.3, "diameter_m": 10.3,
        "notes": "Classic 3-stage parallel burn; boosters + core ignite at liftoff",
        "stages": [
            ("Block B/V/G/D (x4)","RD-107A",  1,  39600,  3784, "4 strap-on boosters; burn with core; mass per booster"),
            ("Block A Core",  "RD-108A",       1,  90360,  6545, "Burns simultaneously with boosters from liftoff"),
            ("Block I (S3)",  "RD-0124",       4,  25400,  2355, "4-chamber upper stage; high Isp"),
        ],
    },
    {
        "name": "Saturn V", "country": "USA", "operator": "NASA",
        "first_flight": 1967, "status": "Retired 1973",
        "payload_leo_kg": 130000, "payload_gto_kg": None, "payload_tli_kg": 48600,
        "glow_kg": 2970000, "height_m": 110.6, "diameter_m": 10.1,
        "notes": "Apollo program; still the most powerful rocket ever flown",
        "stages": [
            ("S-IC First",   "F-1",            5,2077000,130000, "LOX/RP-1; 5 F-1 engines"),
            ("S-II Second",  "J-2",            5, 426000, 36200, "LOX/LH2; 5 J-2 engines"),
            ("S-IVB Third",  "J-2",            1, 106000, 10000, "LOX/LH2; 1 J-2; also used as TMI stage"),
        ],
    },
    {
        "name": "SLS Block 1", "country": "USA", "operator": "NASA",
        "first_flight": 2022, "status": "Active",
        "payload_leo_kg": 95000, "payload_gto_kg": None, "payload_tli_kg": 27000,
        "glow_kg": 2608000, "height_m": 98.0, "diameter_m": 8.4,
        "notes": "Artemis program; RS-25 core + 2 SLS SRBs + ICPS upper stage",
        "stages": [
            ("SLS SRB (x2)",  "SRB (SLS 5-seg)",1,590000, 90718, "5-segment solid strap-on; mass per booster"),
            ("Core Stage",    "RS-25 (SSME)",  4, 987000, 85000, "LOX/LH2; 4 RS-25D engines"),
            ("ICPS Upper",    "RL-10C-3",      1,  27220,  3500, "Interim Cryogenic Propulsion Stage; LOX/LH2"),
        ],
    },
    {
        "name": "New Glenn (7x2 Baseline)", "country": "USA", "operator": "Blue Origin",
        "first_flight": 2025, "status": "Active",
        "payload_leo_kg": 45000, "payload_gto_kg": 13000, "payload_tli_kg": None,
        "glow_kg": 1000000, "height_m": 98.0, "diameter_m": 7.0,
        "notes": "First flight Jan 2025; 7 BE-4 + 2 BE-3U; 7m fairing; Blue Origin claimed 45t LEO / 13t GTO",
        "stages": [
            ("S1 First",     "BE-4",           7, 770000, 55000, "LOX/CH4; 7 BE-4 baseline engines; reusable"),
            ("S2 Upper",     "BE-3U",          2,  75000, 10000, "LOX/LH2; 2 BE-3U baseline engines"),
        ],
    },
    {
        "name": "New Glenn Block 2 (7x2 Upgraded)", "country": "USA", "operator": "Blue Origin",
        "first_flight": 2026, "status": "Development",
        "payload_leo_kg": None, "payload_gto_kg": None, "payload_tli_kg": None,
        "glow_kg": 1000000, "height_m": 98.0, "diameter_m": 7.0,
        "notes": "Upgraded BE-4 Block 2 (2,847 kN SL w/ subcooling) + BE-3U Block 2 (890 kN); no updated payload figures released by Blue Origin",
        "stages": [
            ("S1 First",     "BE-4 Block 2",   7, 770000, 55000, "LOX/CH4; 7 BE-4 Block 2; 640,000 lbf each with subcooling; reusable"),
            ("S2 Upper",     "BE-3U Block 2",  2,  75000, 10000, "LOX/LH2; 2 BE-3U Block 2; 200,000 lbf each"),
        ],
    },
    {
        "name": "New Glenn 9x4", "country": "USA", "operator": "Blue Origin",
        "first_flight": 2027, "status": "Development",
        "payload_leo_kg": 70000, "payload_gto_kg": 14000, "payload_tli_kg": 20000,
        "glow_kg": 1300000, "height_m": 105.0, "diameter_m": 8.7,
        "notes": "Heavy-lift variant; 9 BE-4 Block 2 S1 + 4 BE-3U Block 2 S2; 8.7m fairing; >70t LEO / >14t GTO announced; prop masses estimated",
        "stages": [
            ("S1 First",     "BE-4 Block 2",   9, 990000, 70000, "LOX/CH4; 9 BE-4 Block 2; prop mass estimated ~990t; reusable"),
            ("S2 Upper",     "BE-3U Block 2",  4,  95000, 13000, "LOX/LH2; 4 BE-3U Block 2; prop mass estimated ~95t"),
        ],
    },
    {
        "name": "Starship (IFT-6 config)", "country": "USA", "operator": "SpaceX",
        "first_flight": 2023, "status": "Development",
        "payload_leo_kg": 150000, "payload_gto_kg": None, "payload_tli_kg": None,
        "glow_kg": 5000000, "height_m": 121.0, "diameter_m": 9.0,
        "notes": "Fully reusable target; 150t LEO is aspirational with full reuse; ~200t expendable claimed",
        "stages": [
            ("Super Heavy",  "Raptor 2",      33,3400000,200000, "LOX/CH4; 33 Raptor engines; reusable booster"),
            ("Starship",     "Raptor 2",       6,1200000,100000, "LOX/CH4; 3 vac + 3 SL Raptors; reusable ship"),
        ],
    },
    {
        "name": "Vulcan Centaur", "country": "USA", "operator": "ULA",
        "first_flight": 2024, "status": "Active",
        "payload_leo_kg": 27200, "payload_gto_kg": 9250, "payload_tli_kg": None,
        "glow_kg": 547000, "height_m": 61.6, "diameter_m": 5.4,
        "notes": "Replaces Atlas V and Delta IV; 2 or 4 GEM-63XL strap-ons optional",
        "stages": [
            ("First Stage",  "BE-4",           2, 406000, 22400, "LOX/CH4; 2 BE-4 engines"),
            ("Centaur V",    "RL-10C-1",       2,  54000,  6300, "LOX/LH2; 1 or 2 RL-10C-1; dual-engine shown"),
        ],
    },
]

# ── Propellant Properties ──────────────────────────────────────────────────────
# name, oxidizer, fuel, ox_density_kg_m3, fuel_density_kg_m3, MR_ox_fuel,
# typical_isp_vac_s, typical_isp_sl_s, storable, notes
PROPELLANTS = [
    ("LOX/RP-1",      "LOX",  "RP-1 (Kerosene)", 1141,  820, 2.56, 330, 290, False, "Dense; good SL performance; not storable; workhorse combo"),
    ("LOX/LH2",       "LOX",  "LH2 (Hydrogen)",  1141,   71, 5.50, 455, 380, False, "Highest Isp; low density challenges; large tanks required"),
    ("LOX/CH4",       "LOX",  "CH4 (Methane)",   1141,  423, 3.55, 360, 330, False, "Good Isp + density; reusability-friendly; ISRU potential"),
    ("NTO/MMH",       "NTO",  "MMH",             1450,  880, 1.65, 320, None, True, "Hypergolic; storable; common for upper stages and RCS"),
    ("NTO/UDMH",      "NTO",  "UDMH",            1450,  793, 2.60, 310, None, True, "Hypergolic; storable; Russian/Chinese first stages"),
    ("NTO/Aerozine-50","NTO", "Aerozine-50",     1450,  900, 1.90, 320, None, True, "50/50 N2H4+UDMH; Titan II/IV, Apollo CSM SPS"),
    ("Solid HTPB",    "AP",   "HTPB Binder",     1790, 1050, 3.50, 285, 270, True,  "High density; simple ops; Isp 265-295s depending on formulation"),
    ("Solid PBAN",    "AP",   "PBAN Binder",     1790, 1000, 3.50, 270, 250, True,  "Shuttle/SLS SRB propellant; similar to HTPB"),
    ("H2O2/RP-1",     "H2O2","RP-1",             1440,  820, 7.00, 300, 270, False, "Green propellant; lower Isp than LOX/RP-1; non-cryogenic oxidizer"),
    ("LOX/Ethanol",   "LOX",  "Ethanol",         1141,  789, 1.30, 311, 260, False, "Historical (V-2, Redstone); largely superseded by RP-1"),
]

# ── Subsystem Mass Fraction Model ──────────────────────────────────────────────
# Fractions are expressed as % of propellant mass, per propellant class
# Derived from analysis of real vehicles (Falcon 9, Atlas V, Saturn V, Ariane 5, Soyuz)
# Engine mass is handled separately (from engine DB x engine count)
SUBSYSTEMS = [
    # subsystem, LOX/RP1%, LOX/LH2%, LOX/CH4%, Solid%, Hypergolic%, description
    ("Tank Structure",         3.5, 6.5, 3.8, 1.8, 2.5, "Propellant tanks, interstage, skirt structure; LH2 needs insulation"),
    ("Pressurization System",  0.4, 0.6, 0.4, 0.0, 0.5, "He pressurant tanks, regulators, lines; N/A for solids"),
    ("Feed System / Plumbing", 0.8, 1.0, 0.8, 0.2, 0.6, "Turbopumps, manifolds, valves, lines (exclusive of engine)"),
    ("Avionics / GNC",         0.3, 0.4, 0.3, 0.2, 0.4, "Flight computer, IMU, GPS, telemetry; typically fixed-ish mass"),
    ("Power System",           0.2, 0.3, 0.2, 0.1, 0.2, "Batteries, power distribution, wiring harness"),
    ("Thermal Control",        0.2, 1.0, 0.3, 0.1, 0.3, "MLI blankets, heaters, paint; LH2 stages significantly higher"),
    ("RCS / ACS",              0.3, 0.4, 0.3, 0.1, 0.2, "Reaction control thrusters, propellant, lines"),
    ("Separation System",      0.1, 0.1, 0.1, 0.1, 0.1, "Stage separation clampband / ordnance; fairing sep if applicable"),
    ("Recovery Hardware",      0.0, 0.0, 0.0, 0.0, 0.0, "Landing legs, grid fins, parachutes — add manually if reusable"),
    ("Residuals / Margin",     0.5, 0.8, 0.5, 0.3, 0.5, "Unusable propellant, trapped fluids, 3-5% mass growth margin"),
]


def style_hdr(ws, row, col, text, span=1, dark=True):
    cell = ws.cell(row=row, column=col, value=text)
    bg = C["hdr"] if dark else C["sub"]
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell


def style_sec(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=C["sec"])
    cell.font = Font(bold=True, size=10, color="1F1F1F")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell


def style_input(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=C["input"])
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def style_calc(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=C["calc"])
    cell.font = Font(size=10, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def style_label(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=C["ltgray"])
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return cell


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    b = thin_border()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = b


def col(n):
    return get_column_letter(n)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Engine Database
# ─────────────────────────────────────────────────────────────────────────────
def build_engine_db(wb):
    ws = wb.create_sheet("Engine DB")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = [
        "Engine Name", "Vehicle Application", "Propellant", "Vac Isp (s)", "SL Isp (s)",
        "Vac Thrust (kN)", "SL Thrust (kN)", "Engine Mass (kg)", "Chamber Press (bar)",
        "Expansion Ratio", "Notes"
    ]
    widths = [18, 22, 12, 12, 12, 16, 16, 16, 18, 15, 45]

    style_hdr(ws, 1, 1, "ROCKET ENGINE DATABASE — Real-World Performance Data", span=len(headers))
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_hdr(ws, 2, ci, h, dark=False)
        ws.column_dimensions[col(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    # Group by propellant
    groups = {}
    for eng in ENGINES:
        groups.setdefault(eng[2], []).append(eng)

    r = 3
    for prop_group, engines in groups.items():
        style_sec(ws, r, 1, f"── {prop_group} ──", span=len(headers))
        ws.row_dimensions[r].height = 18
        r += 1
        for eng in engines:
            for ci, val in enumerate(eng, 1):
                cell = ws.cell(r, ci, value=val if val is not None else "—")
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center" if ci > 2 else "left",
                                           vertical="center", wrap_text=(ci == 11))
                if ci in (1, 2, 11):
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               wrap_text=True, indent=1)
            apply_border_range(ws, r, r, 1, len(headers))
            ws.row_dimensions[r].height = 15
            r += 1

    ws.auto_filter.ref = f"A2:{col(len(headers))}2"
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Vehicle Database
# ─────────────────────────────────────────────────────────────────────────────
def build_vehicle_db(wb):
    ws = wb.create_sheet("Vehicle DB")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    style_hdr(ws, 1, 1, "REFERENCE LAUNCH VEHICLE DATABASE — Real Vehicle Mass & Performance Data", span=12)

    # Vehicle summary headers
    vh = ["Vehicle", "Country", "Operator", "First Flight", "Status",
          "Payload LEO (kg)", "Payload GTO (kg)", "Payload TLI (kg)",
          "GLOW (kg)", "Height (m)", "Diameter (m)", "Notes"]
    vw = [22, 10, 14, 12, 14, 16, 16, 16, 14, 11, 12, 45]
    for ci, (h, w) in enumerate(zip(vh, vw), 1):
        style_hdr(ws, 2, ci, h, dark=False)
        ws.column_dimensions[col(ci)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    r = 3
    for v in VEHICLES:
        row_vals = [v["name"], v["country"], v["operator"], v["first_flight"],
                    v["status"], v["payload_leo_kg"], v["payload_gto_kg"],
                    v.get("payload_tli_kg", None), v["glow_kg"],
                    v["height_m"], v["diameter_m"], v["notes"]]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(r, ci, value=val if val is not None else "—")
            cell.font = Font(bold=(ci == 1), size=10)
            cell.alignment = Alignment(horizontal="left" if ci in (1,2,3,5,12) else "center",
                                       vertical="center", wrap_text=(ci == 12), indent=1)
            cell.fill = PatternFill("solid", fgColor=C["ltgray"] if r % 2 == 0 else C["white"])
        apply_border_range(ws, r, r, 1, len(vh))
        ws.row_dimensions[r].height = 15
        r += 1

    # Stage breakdown section
    r += 1
    style_hdr(ws, r, 1, "PER-STAGE MASS BREAKDOWN", span=9)
    r += 1
    sh = ["Vehicle", "Stage", "Engine", "# Engines", "Propellant Mass (kg)",
          "Dry Mass (kg)", "Gross Mass (kg)", "Struct. Frac. (%)", "Notes"]
    sw = [22, 20, 18, 10, 20, 15, 15, 15, 45]
    for ci, (h, w) in enumerate(zip(sh, sw), 1):
        style_hdr(ws, r, ci, h, dark=False)
    ws.row_dimensions[r].height = 25
    r += 1

    for v in VEHICLES:
        for si, stg in enumerate(v["stages"]):
            sname, eng, n_eng, prop_mass, dry_mass, notes = stg
            gross = prop_mass + dry_mass
            sf = (dry_mass - 0) / prop_mass * 100  # simplified: dry/prop
            row_vals = [v["name"] if si == 0 else "", sname, eng, n_eng,
                        prop_mass, dry_mass, gross, round(sf, 2), notes]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(r, ci, value=val)
                cell.font = Font(bold=(ci == 1 and si == 0), size=10)
                cell.alignment = Alignment(horizontal="left" if ci in (1,2,3,9) else "center",
                                           vertical="center", wrap_text=(ci == 9), indent=1)
                cell.fill = PatternFill("solid", fgColor=C["ltgray"] if r % 2 == 0 else C["white"])
            apply_border_range(ws, r, r, 1, len(sh))
            ws.row_dimensions[r].height = 15
            r += 1

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Propellants
# ─────────────────────────────────────────────────────────────────────────────
def build_propellants(wb):
    ws = wb.create_sheet("Propellants")
    ws.sheet_view.showGridLines = False

    style_hdr(ws, 1, 1, "PROPELLANT PROPERTIES", span=10)

    headers = ["Propellant", "Oxidizer", "Fuel", "Ox Density (kg/m³)",
               "Fuel Density (kg/m³)", "Mix Ratio (O/F)", "Bulk Density* (kg/m³)",
               "Typical Vac Isp (s)", "Storable?", "Notes"]
    widths = [18, 12, 18, 18, 18, 15, 18, 18, 10, 50]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_hdr(ws, 2, ci, h, dark=False)
        ws.column_dimensions[col(ci)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    for ri, p in enumerate(PROPELLANTS, 3):
        name, ox, fuel, ox_d, f_d, mr, isp_vac, isp_sl, storable, notes = p
        bulk = (ox_d * mr + f_d * 1) / (mr + 1)
        row_vals = [name, ox, fuel, ox_d, f_d, mr, round(bulk, 0),
                    isp_vac, "Yes" if storable else "No", notes]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(ri, ci, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left" if ci in (1,2,3,9,10) else "center",
                                       vertical="center", wrap_text=(ci == 10), indent=1)
            cell.fill = PatternFill("solid", fgColor=C["ltgray"] if ri % 2 == 0 else C["white"])
        apply_border_range(ws, ri, ri, 1, len(headers))
        ws.row_dimensions[ri].height = 15

    r = len(PROPELLANTS) + 4
    ws.cell(r, 1, value="* Bulk density = mass-weighted average of oxidizer + fuel at stated mixture ratio").font = Font(italic=True, size=9, color="555555")

    # Subsystem mass fraction model
    r += 2
    style_hdr(ws, r, 1, "SUBSYSTEM MASS FRACTION MODEL — Fractions as % of Propellant Mass", span=7)
    r += 1
    sh = ["Subsystem", "LOX/RP-1 (%)", "LOX/LH2 (%)", "LOX/CH4 (%)",
          "Solid (%)", "Hypergolic (%)", "Description"]
    sw2 = [22, 13, 13, 13, 10, 14, 55]
    for ci, (h, w) in enumerate(zip(sh, sw2), 1):
        style_hdr(ws, r, ci, h, dark=False)
        ws.column_dimensions[col(ci)].width = max(ws.column_dimensions[col(ci)].width, w)
    r += 1

    totals = [0.0] * 5
    for sub in SUBSYSTEMS:
        name_s, rp1, lh2, ch4, sol, hyp, desc = sub
        for ci, val in enumerate([name_s, rp1, lh2, ch4, sol, hyp, desc], 1):
            cell = ws.cell(r, ci, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left" if ci in (1, 7) else "center",
                                       vertical="center", wrap_text=(ci == 7), indent=1)
            cell.fill = PatternFill("solid", fgColor=C["ltgray"] if r % 2 == 0 else C["white"])
        apply_border_range(ws, r, r, 1, 7)
        ws.row_dimensions[r].height = 15
        totals = [totals[i] + v for i, v in enumerate([rp1, lh2, ch4, sol, hyp])]
        r += 1

    # Totals row
    ws.cell(r, 1, value="TOTAL (excl. engines)").font = Font(bold=True, size=10)
    for ci, t in enumerate(totals, 2):
        cell = ws.cell(r, ci, value=round(t, 2))
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor=C["sec"])
        cell.alignment = Alignment(horizontal="center")
    apply_border_range(ws, r, r, 1, 7)

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Vehicle Design (main working sheet)
# ─────────────────────────────────────────────────────────────────────────────
def build_design(wb):
    ws = wb.create_sheet("Vehicle Design", 0)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    # Column layout
    # A: labels   B: value/unit   C: (blank buffer)
    # D..J: Stage 1-3 (three cols each: label, value, spacer)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 4
    for i in range(MAX_STAGES):
        ws.column_dimensions[col(4 + i * 3)].width = 18
        ws.column_dimensions[col(5 + i * 3)].width = 16
        ws.column_dimensions[col(6 + i * 3)].width = 3

    # Row 1: Title
    style_hdr(ws, 1, 1, "LAUNCH VEHICLE DESIGN TOOL", span=4 + MAX_STAGES * 3)
    ws.row_dimensions[1].height = 28

    # Row 2: Legend
    ws.merge_cells(f"A2:{col(4 + MAX_STAGES * 3)}2")
    ws.cell(2, 1, "Yellow = Input   |   Blue = Calculated   |   "
                  "Engine name must match Engine DB exactly   |   "
                  "ΔV Allocation per stage must sum to Mission ΔV").font = Font(size=9, italic=True)

    # Row 3: Stage column headers
    ws.cell(3, 1, value="PARAMETER").font = Font(bold=True, size=10)
    ws.cell(3, 2, value="UNITS").font = Font(bold=True, size=10)
    ws.cell(3, 1).fill = PatternFill("solid", fgColor=C["sec"])
    ws.cell(3, 2).fill = PatternFill("solid", fgColor=C["sec"])
    ws.row_dimensions[3].height = 20

    for si in range(MAX_STAGES):
        c_label = 4 + si * 3
        c_val   = 5 + si * 3
        ws.merge_cells(start_row=3, start_column=c_label,
                       end_row=3, end_column=c_val)
        cell = ws.cell(3, c_label, value=f"Stage {si+1}")
        cell.fill = PatternFill("solid", fgColor=C["stage"][si])
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Section: Mission ──────────────────────────────────────────────────────
    r = 5
    style_sec(ws, r, 1, "MISSION REQUIREMENTS", span=2)
    r += 1

    mission_inputs = [
        ("Vehicle Name",            "text",  "New Glenn Block 2 (7x2)",  None),
        ("Target Orbit",            "text",  "LEO 400km",                None),
        ("Payload Adapter Mass",    "kg",    200,          "Separation hardware mass between payload and vehicle"),
        ("Orbital Velocity (LEO)",  "m/s",   7784,         "Target orbital velocity — 7784 m/s for 400km LEO; 7726 m/s for 200km"),
        ("Drag Loss",               "m/s",   120,          "Aerodynamic drag loss — typically 100-150 m/s; relatively insensitive to thrust"),
        ("Gravity Loss Coefficient","—",      1750,         "Gravity loss = coeff / TWR_liftoff; calibrated to ~1450 m/s at TWR=1.2; adjust to match known vehicles"),
        ("GTO Orbital ΔV",          "m/s",   10400,        "Velocity needed at GTO perigee — ~10,200-10,500 m/s depending on inclination and target apogee"),
        ("Number of Stages",        "1-3",   2,            "Active stages (enter 1-3)"),
    ]

    mission_rows = {}
    for label, unit, default, note in mission_inputs:
        ws.cell(r, 1, value=label).font = Font(size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
        style_input(ws, r, 2, default)
        ws.cell(r, 2).number_format = "#,##0" if unit in ("kg", "m/s") else "@"
        if note:
            ws.cell(r, 3, value=f"← {note}").font = Font(size=8, italic=True, color="777777")
            ws.merge_cells(start_row=r, start_column=3,
                           end_row=r, end_column=4 + MAX_STAGES * 3 - 1)
        mission_rows[label] = r
        r += 1

    adapter_row        = mission_rows["Payload Adapter Mass"]
    orb_vel_row        = mission_rows["Orbital Velocity (LEO)"]
    drag_loss_row      = mission_rows["Drag Loss"]
    grav_coeff_row     = mission_rows["Gravity Loss Coefficient"]
    gto_orb_dv_row     = mission_rows["GTO Orbital ΔV"]
    n_stages_row       = mission_rows["Number of Stages"]

    # LEO and GTO mission ΔV will be calculated after stage config (needs TWR)
    # placeholders assigned after stage config loop
    dv_budget_row     = None  # set later
    dv_gto_budget_row = None  # set later

    # ── Section: Stage Configuration ─────────────────────────────────────────
    r += 1
    style_sec(ws, r, 1, "STAGE CONFIGURATION", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    stage_rows = {}
    stage_params = [
        ("Stage Name",              "text",  ["S1 First",    "S2 Upper",     ""]),
        ("Engine (from Engine DB)", "name",  ["BE-4 Block 2","BE-3U Block 2",""]),
        ("Number of Engines",       "count", [7,              2,              ""]),
        ("Propellant Combination",  "type",  ["LOX/CH4",      "LOX/LH2",      ""]),
        ("Propellant Mass (nominal)","kg",    [770000,          75000,          ""]),
        ("Subcooling Gain (%)",      "%",     [3.5,             0.0,            0.0]),
        ("Effective Propellant Mass","kg",    None),   # nominal × (1 + gain/100)
        ("LEO ΔV Fraction",          "0-1",  [0.597,       0.403,     ""]),
        ("LEO ΔV Allocation",        "m/s",  None),   # fraction × calculated LEO mission ΔV
        ("Vac Isp Override",         "s",    ["",          "",        ""]),
        ("Vac Isp (from DB)",        "s",    None),   # VLOOKUP col 4
        ("SL Isp (from DB)",         "s",    None),   # VLOOKUP col 5
        ("Atm Fraction (0-1)",       "—",    [0.6,         1.0,       1.0]),
        ("Effective Isp (used)",     "s",    None),   # SL + frac*(Vac-SL), or override
        ("GTO ΔV Fraction",          "0-1",  [0.555,       0.445,     ""]),
        ("GTO ΔV Allocation",        "m/s",  None),   # fraction × calculated GTO mission ΔV
    ]

    for param, unit, defaults in stage_params:
        ws.cell(r, 1, value=param).font = Font(size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
        ws.cell(r, 2, value=unit).font = Font(size=9, color="777777")
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
        stage_rows[param] = r

        for si in range(MAX_STAGES):
            c_val = 5 + si * 3
            if param in ("LEO ΔV Allocation", "GTO ΔV Allocation"):
                # Formula filled after dv_budget_row / dv_gto_budget_row is known (deferred)
                frac_key = "LEO ΔV Fraction" if param == "LEO ΔV Allocation" else "GTO ΔV Fraction"
                frac_c = f"{col(c_val)}{stage_rows[frac_key]}"
                cell = style_calc(ws, r, c_val, value="")
                cell.number_format = "#,##0"
                if not hasattr(ws, '_dv_alloc_deferred'):
                    ws._dv_alloc_deferred = []
                ws._dv_alloc_deferred.append((r, c_val, frac_c, param))
            elif param == "Effective Propellant Mass":
                nom_c  = f"{col(c_val)}{stage_rows['Propellant Mass (nominal)']}"
                gain_c = f"{col(c_val)}{stage_rows['Subcooling Gain (%)']}"
                formula = f'=IFERROR({nom_c}*(1+{gain_c}/100),"")'
                cell = style_calc(ws, r, c_val, value=formula)
                cell.font = Font(bold=True, size=10, italic=True)
                cell.number_format = "#,##0"
            elif param == "Vac Isp (from DB)":
                eng_cell = f"{col(c_val)}{stage_rows['Engine (from Engine DB)']}"
                ovr_cell = f"{col(c_val)}{stage_rows['Vac Isp Override']}"
                formula = (f'=IF({ovr_cell}<>"",{ovr_cell},'
                           f'IFERROR(VLOOKUP({eng_cell},\'Engine DB\'!$A:$K,4,0),"Enter engine"))')
                cell = style_calc(ws, r, c_val, value=formula)
                cell.number_format = "0.0"
            elif param == "SL Isp (from DB)":
                eng_cell = f"{col(c_val)}{stage_rows['Engine (from Engine DB)']}"
                # Col 5 = SL Isp; returns "—" if not applicable (upper stages, solids)
                formula = f'=IFERROR(VLOOKUP({eng_cell},\'Engine DB\'!$A:$K,5,0),"—")'
                cell = style_calc(ws, r, c_val, value=formula)
                cell.number_format = "0.0"
            elif param == "Effective Isp (used)":
                vac_c  = f"{col(c_val)}{stage_rows['Vac Isp (from DB)']}"
                sl_c   = f"{col(c_val)}{stage_rows['SL Isp (from DB)']}"
                frac_c = f"{col(c_val)}{stage_rows['Atm Fraction (0-1)']}"
                # If SL Isp unavailable (solid/upper), use Vac Isp directly
                formula = (f'=IFERROR('
                           f'IF(ISNUMBER({sl_c}),'
                           f'{sl_c}+{frac_c}*({vac_c}-{sl_c}),'
                           f'{vac_c}),{vac_c})')
                cell = style_calc(ws, r, c_val, value=formula)
                cell.font = Font(bold=True, size=10, italic=True)
                cell.number_format = "0.0"
            else:
                val = defaults[si] if si < len(defaults) else ""
                cell = style_input(ws, r, c_val, val)
                cell.number_format = "#,##0" if unit in ("kg", "m/s") else "@"
            cell.fill = PatternFill("solid", fgColor=C["calc"] if defaults is None or param in ("Vac Isp (from DB)", "SL Isp (from DB)", "Effective Isp (used)", "Effective Propellant Mass", "LEO ΔV Allocation", "GTO ΔV Allocation") else C["input"])

        r += 1

    # LEO ΔV allocation check row
    dv_alloc_row = stage_rows["LEO ΔV Allocation"]

    # ── Mission ΔV Analysis (calculated from TWR) ─────────────────────────────
    # TWR uses sum of all stage gross masses (prop + dry) as GLOW approximation.
    # This avoids a circular reference with payload. Error is small (<3%) since
    # payload is ~2-4% of GLOW. Stage 1 SL thrust used for gravity loss calculation.
    r += 1
    style_sec(ws, r, 1, "MISSION ΔV ANALYSIS  —  Gravity loss computed from liftoff TWR", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    # Stage gross masses (prop nominal + dry) — needed for GLOW estimate
    # We use nominal prop mass here (before subcooling) as that's the tank-full condition at liftoff
    nom_prop_row = stage_rows["Propellant Mass (nominal)"]

    def make_label_calc(ws, r, label, unit):
        ws.cell(r, 1, value=label).font = Font(size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
        ws.cell(r, 2, value=unit).font = Font(size=9, color="777777")
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
        ws.cell(r, 2).alignment = Alignment(horizontal="center")

    # Estimated GLOW (sum of all stage prop+dry, no payload — conservative TWR)
    make_label_calc(ws, r, "Est. GLOW (stage masses only, no payload)", "kg")
    glow_est_row = r
    # We need eff_dry_row but it's not defined yet — use a forward reference string
    # We'll fill this formula after eff_dry_row is assigned, so store the row and come back
    glow_est_r = r  # store row index; formula filled after subsystem section
    r += 1

    # Stage 1 liftoff TWR
    make_label_calc(ws, r, "Stage 1 Liftoff TWR (est., no payload)", "—")
    twr_est_row = r
    twr_est_r = r
    r += 1

    # Gravity loss = coeff / TWR
    make_label_calc(ws, r, "Gravity Loss (= coeff / TWR)", "m/s")
    grav_loss_row = r
    ws.cell(r, 2, value=f"=IFERROR(B{grav_coeff_row}/B{twr_est_row},\"\")").font = Font(size=10, italic=True)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["calc"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    ws.cell(r, 2).number_format = "#,##0"
    r += 1

    # LEO Mission ΔV = orbital velocity + drag + gravity loss
    make_label_calc(ws, r, "LEO Mission ΔV (calculated)", "m/s")
    leo_dv_calc_row = r
    ws.cell(r, 2, value=f"=IFERROR(B{orb_vel_row}+B{drag_loss_row}+B{grav_loss_row},\"\")").font = Font(bold=True, size=10, italic=True)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["calc"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    ws.cell(r, 2).number_format = "#,##0"
    dv_budget_row = leo_dv_calc_row
    r += 1

    # GTO Mission ΔV = GTO orbital ΔV + drag + gravity loss
    # GTO gravity loss is slightly lower (higher energy trajectory) but we use same estimate
    make_label_calc(ws, r, "GTO Mission ΔV (calculated)", "m/s")
    gto_dv_calc_row = r
    ws.cell(r, 2, value=f"=IFERROR(B{gto_orb_dv_row}+B{drag_loss_row}+B{grav_loss_row},\"\")").font = Font(bold=True, size=10, italic=True)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["calc"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    ws.cell(r, 2).number_format = "#,##0"
    dv_gto_budget_row = gto_dv_calc_row
    r += 1

    # LEO ΔV allocation check
    alloc_sum = "+".join(f"IFERROR({col(5+si*3)}{dv_alloc_row}*1,0)" for si in range(MAX_STAGES))
    ws.cell(r, 1, value="LEO ΔV Allocation Check (should = LEO Mission ΔV above)").font = Font(size=10, italic=True)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value=f"={alloc_sum}").font = Font(bold=True, size=10)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["calc"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    ws.cell(r, 2).number_format = "#,##0"
    ws.cell(r, 3, value=f'=IF(ABS(B{r}-B{dv_budget_row})<1,"✓ OK","⚠ Adjust stage allocations to sum to LEO Mission ΔV")').font = Font(size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4 + MAX_STAGES * 3 - 1)
    r += 1

    # ── Section: Engine Mass (auto-looked-up) ─────────────────────────────────
    ws.cell(r, 1, value="Engine Mass per Engine").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    engine_mass_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        eng_cell = f"{col(c_val)}{stage_rows['Engine (from Engine DB)']}"
        formula = f'=IFERROR(VLOOKUP({eng_cell},\'Engine DB\'!$A:$K,8,0),"—")'
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # ── Section: Subsystem Mass Breakdown ────────────────────────────────────
    r += 1
    style_sec(ws, r, 1, "SUBSYSTEM MASS BREAKDOWN (auto-estimated; override any yellow cell)", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    prop_row           = stage_rows["Propellant Combination"]
    base_prop_mass_row = stage_rows["Propellant Mass (nominal)"]  # used for subsystem fractions
    prop_mass_row      = stage_rows["Effective Propellant Mass"]   # used for all performance calcs

    subsys_rows = {}
    for sub in SUBSYSTEMS:
        sub_name, rp1, lh2, ch4, sol, hyp, desc = sub
        ws.cell(r, 1, value=sub_name).font = Font(size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=2)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
        ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
        subsys_rows[sub_name] = r

        for si in range(MAX_STAGES):
            c_val = 5 + si * 3
            prop_cell  = f"{col(c_val)}{prop_row}"
            pmass_cell = f"{col(c_val)}{base_prop_mass_row}"
            formula = (
                f'=IF({pmass_cell}="","",{pmass_cell}/100*'
                f'IF({prop_cell}="LOX/RP-1",{rp1},'
                f'IF({prop_cell}="LOX/LH2",{lh2},'
                f'IF({prop_cell}="LOX/CH4",{ch4},'
                f'IF(OR({prop_cell}="Solid HTPB",{prop_cell}="Solid PBAN"),{sol},'
                f'{hyp})))))'
            )
            cell = style_input(ws, r, c_val, value=formula)
            cell.fill = PatternFill("solid", fgColor=C["input"])
            cell.number_format = "#,##0"
        r += 1

    # Engine system mass (engine mass × count × 1.15 for feed system hardware)
    ws.cell(r, 1, value="Engine System Total").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=2)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    engine_sys_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        n_eng_cell    = f"{col(c_val)}{stage_rows['Number of Engines']}"
        eng_mass_cell = f"{col(c_val)}{engine_mass_row}"
        formula = f'=IFERROR({n_eng_cell}*{eng_mass_cell}*1.15, "")'
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # Total estimated dry mass
    ws.cell(r, 1, value="TOTAL DRY MASS (Estimated)").font = Font(bold=True, size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["sec"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["sec"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    dry_mass_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        sub_refs = "+".join(f"{col(c_val)}{subsys_rows[s[0]]}" for s in SUBSYSTEMS)
        eng_ref  = f"{col(c_val)}{engine_sys_row}"
        formula  = f"=IFERROR({sub_refs}+{eng_ref}, \"\")"
        cell = style_calc(ws, r, c_val, value=formula)
        cell.font = Font(bold=True, size=10, italic=True)
        cell.number_format = "#,##0"
    r += 1

    # Dry mass override (enter known value to replace estimate)
    ws.cell(r, 1, value="Dry Mass Override (enter known value)").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    dry_override_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        style_input(ws, r, c_val, value="")
        ws.cell(r, c_val).number_format = "#,##0"
    r += 1

    # Effective dry mass: override if present, else estimate
    ws.cell(r, 1, value="Effective Dry Mass").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    eff_dry_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        ov = f"{col(c_val)}{dry_override_row}"
        es = f"{col(c_val)}{dry_mass_row}"
        formula = f'=IF({ov}<>"",{ov},{es})'
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # ── Fill deferred GLOW and TWR formulas (needed eff_dry_row) ─────────────
    # GLOW = sum of all stages (nominal prop + effective dry) — no payload, conservative TWR
    glow_parts = "+".join(
        f"IFERROR({col(5+si*3)}{nom_prop_row}+{col(5+si*3)}{eff_dry_row},0)"
        for si in range(MAX_STAGES)
    )
    glow_cell = ws.cell(glow_est_r, 2, value=f"={glow_parts}")
    glow_cell.font = Font(size=10, italic=True)
    glow_cell.fill = PatternFill("solid", fgColor=C["calc"])
    glow_cell.alignment = Alignment(horizontal="center")
    glow_cell.number_format = "#,##0"

    # TWR = Stage 1 SL thrust (kN→N) × n_engines / (GLOW × g0)
    s1_eng_cell  = f"{col(5)}{stage_rows['Engine (from Engine DB)']}"
    s1_neng_cell = f"{col(5)}{stage_rows['Number of Engines']}"
    sl_thrust_f  = f"IFERROR(VLOOKUP({s1_eng_cell},'Engine DB'!$A:$K,7,0),0)"
    twr_formula  = f"=IFERROR({s1_neng_cell}*{sl_thrust_f}*1000/(B{glow_est_r}*{G0}),\"\")"
    twr_cell = ws.cell(twr_est_r, 2, value=twr_formula)
    twr_cell.font = Font(bold=True, size=10, italic=True)
    twr_cell.fill = PatternFill("solid", fgColor=C["calc"])
    twr_cell.alignment = Alignment(horizontal="center")
    twr_cell.number_format = "0.00"

    # ── Fill deferred ΔV Allocation formulas (needed dv_budget_row / dv_gto_budget_row) ──
    for (row_r, c_val, frac_c, param) in getattr(ws, '_dv_alloc_deferred', []):
        budget = dv_budget_row if param == "LEO ΔV Allocation" else dv_gto_budget_row
        formula = f'=IFERROR({frac_c}*B{budget},"")'
        cell = ws.cell(row_r, c_val, value=formula)
        cell.fill = PatternFill("solid", fgColor=C["calc"])
        cell.font = Font(size=10, italic=True)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "#,##0"

    # ── Section: Performance ─────────────────────────────────────────────────
    r += 1
    style_sec(ws, r, 1, "PERFORMANCE CALCULATIONS", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Max Payload (primary output) ──────────────────────────────────────────
    # Derived analytically from the last active stage using Tsiolkovsky inverted:
    #   m_above = prop / (EXP(dv_alloc / (Isp * g0)) - 1) - dry_mass
    #   max_payload = m_above - adapter_mass
    # The last active stage is determined by Number of Stages input.
    # We compute for each stage and select the appropriate one.
    ws.cell(r, 1, value="MAX PAYLOAD TO ORBIT").font = Font(bold=True, size=12)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["pos"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["pos"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    max_payload_row = r
    adapter_cell = f"B{adapter_row}"
    n_stages_cell = f"B{n_stages_row}"

    # Build per-stage payload formula: prop/(EXP(dv/(isp*g0))-1) - dry - adapter
    stage_payload_formulas = []
    for si in range(MAX_STAGES):
        c_val    = 5 + si * 3
        prop_c   = f"{col(c_val)}{prop_mass_row}"
        isp_c    = f"{col(c_val)}{stage_rows['Effective Isp (used)']}"
        dv_c     = f"{col(c_val)}{dv_alloc_row}"
        dry_c    = f"{col(c_val)}{eff_dry_row}"
        f = (f"IFERROR({prop_c}/(EXP({dv_c}/({isp_c}*{G0}))-1)"
             f"-{dry_c}-{adapter_cell}, \"\")")
        stage_payload_formulas.append(f)

    # Select payload from last active stage via nested IF on n_stages
    # IF(n=3, stage3_formula, IF(n=2, stage2_formula, stage1_formula))
    nested = stage_payload_formulas[0]
    for si in range(1, MAX_STAGES):
        nested = f"IF({n_stages_cell}>={si+1},{stage_payload_formulas[si]},{nested})"
    payload_formula = f"={nested}"

    cell = style_calc(ws, r, 2, value=payload_formula)
    cell.font = Font(bold=True, size=12, italic=True)
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=C["pos"])
    note_cell = ws.cell(r, 3, value="← Solved from last stage: m_payload = prop/(e^(ΔV/Isp·g₀)−1) − dry − adapter")
    note_cell.font = Font(size=8, italic=True, color="444444")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4 + MAX_STAGES * 3 - 1)
    r += 1

    # ── MAX PAYLOAD TO GTO (analytical — same method as LEO) ─────────────────
    r += 1
    style_sec(ws, r, 1, "GTO PERFORMANCE", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    ws.cell(r, 1, value="MAX PAYLOAD TO GTO").font = Font(bold=True, size=12)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["pos"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["pos"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    max_payload_gto_row = r

    gto_dv_alloc_row = stage_rows["GTO ΔV Allocation"]
    gto_stage_payload_formulas = []
    for si in range(MAX_STAGES):
        c_val  = 5 + si * 3
        prop_c = f"{col(c_val)}{prop_mass_row}"
        isp_c  = f"{col(c_val)}{stage_rows['Effective Isp (used)']}"
        dv_c   = f"{col(c_val)}{gto_dv_alloc_row}"
        dry_c  = f"{col(c_val)}{eff_dry_row}"
        f = (f"IFERROR({prop_c}/(EXP({dv_c}/({isp_c}*{G0}))-1)"
             f"-{dry_c}-{adapter_cell}, \"\")")
        gto_stage_payload_formulas.append(f)

    gto_nested = gto_stage_payload_formulas[0]
    for si in range(1, MAX_STAGES):
        gto_nested = f"IF({n_stages_cell}>={si+1},{gto_stage_payload_formulas[si]},{gto_nested})"

    cell = style_calc(ws, r, 2, value=f"={gto_nested}")
    cell.font = Font(bold=True, size=12, italic=True)
    cell.number_format = "#,##0"
    cell.fill = PatternFill("solid", fgColor=C["pos"])
    note_cell = ws.cell(r, 3, value="← Solved from last stage using GTO ΔV allocations: m_payload = prop/(e^(ΔV/Isp·g₀)−1) − dry − adapter")
    note_cell.font = Font(size=8, italic=True, color="444444")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4 + MAX_STAGES * 3 - 1)
    r += 1

    # ── Stage-by-stage verification (Tsiolkovsky forward check) ──────────────
    # Now that we have max_payload, compute actual ΔV per stage as a check
    # gross_mass(i) = prop(i) + dry(i)
    ws.cell(r, 1, value="Gross Stage Mass (Prop + Dry)").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    gross_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        formula = (f"=IFERROR({col(c_val)}{prop_mass_row}"
                   f"+{col(c_val)}{eff_dry_row}, \"\")")
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # Burnout mass (m_f): dry of this stage + all stages above + max_payload + adapter
    ws.cell(r, 1, value="Burnout Mass m_f").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    burnout_row = r
    payload_calc_cell = f"B{max_payload_row}"
    for si in range(MAX_STAGES):
        c_val    = 5 + si * 3
        dry_cell = f"{col(c_val)}{eff_dry_row}"
        above_gross = "+".join(
            f"IFERROR({col(5 + j * 3)}{gross_row}*1,0)"
            for j in range(si + 1, MAX_STAGES)
        )
        if above_gross:
            formula = (f"=IFERROR({dry_cell}+{above_gross}"
                       f"+{payload_calc_cell}+{adapter_cell}, \"\")")
        else:
            formula = (f"=IFERROR({dry_cell}"
                       f"+{payload_calc_cell}+{adapter_cell}, \"\")")
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # Initial mass m_0: burnout + propellant
    ws.cell(r, 1, value="Initial Mass m_0 (at ignition)").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="kg").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    m0_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        formula = (f"=IFERROR({col(c_val)}{burnout_row}"
                   f"+{col(c_val)}{prop_mass_row}, \"\")")
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "#,##0"
    r += 1

    # Mass ratio
    ws.cell(r, 1, value="Mass Ratio (m₀ / m_f)").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="—").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    mr_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        formula = (f"=IFERROR({col(c_val)}{m0_row}"
                   f"/{col(c_val)}{burnout_row}, \"\")")
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "0.000"
    r += 1

    # Actual Delta-V (verification — should match ΔV Allocation if vehicle is sized correctly)
    ws.cell(r, 1, value="Actual Stage ΔV (verification)").font = Font(bold=True, size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["sec"])
    ws.cell(r, 2, value="m/s").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["sec"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    dv_row = r
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        isp_c = f"{col(c_val)}{stage_rows['Effective Isp (used)']}"
        mr_c  = f"{col(c_val)}{mr_row}"
        formula = f"=IFERROR({isp_c}*{G0}*LN({mr_c}), \"\")"
        cell = style_calc(ws, r, c_val, value=formula)
        cell.font = Font(bold=True, size=10, italic=True)
        cell.number_format = "#,##0"
    r += 1

    # TWR at stage ignition
    ws.cell(r, 1, value="Thrust-to-Weight Ratio (at ignition)").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="—").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    twr_row = r
    for si in range(MAX_STAGES):
        c_val      = 5 + si * 3
        eng_cell   = f"{col(c_val)}{stage_rows['Engine (from Engine DB)']}"
        n_eng_cell = f"{col(c_val)}{stage_rows['Number of Engines']}"
        m0_cell    = f"{col(c_val)}{m0_row}"
        # Stage 1: use sea-level thrust; upper stages: vacuum thrust
        if si == 0:
            thrust_f = f"IFERROR(VLOOKUP({eng_cell},'Engine DB'!$A:$K,7,0),0)"
        else:
            thrust_f = f"IFERROR(VLOOKUP({eng_cell},'Engine DB'!$A:$K,6,0),0)"
        formula = f"=IFERROR({n_eng_cell}*{thrust_f}*1000/({m0_cell}*{G0}), \"\")"
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "0.00"
    r += 1

    # Propellant mass fraction
    ws.cell(r, 1, value="Propellant Mass Fraction").font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2, value="—").font = Font(size=9, color="777777")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["ltgray"])
    ws.cell(r, 2).alignment = Alignment(horizontal="center")
    for si in range(MAX_STAGES):
        c_val = 5 + si * 3
        formula = (f"=IFERROR({col(c_val)}{prop_mass_row}"
                   f"/{col(c_val)}{m0_row}, \"\")")
        cell = style_calc(ws, r, c_val, value=formula)
        cell.number_format = "0.000"
    r += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    r += 1
    style_sec(ws, r, 1, "VEHICLE SUMMARY", span=2 + MAX_STAGES * 3)
    ws.row_dimensions[r].height = 18
    r += 1

    dv_actual_sum = "+".join(
        f"IFERROR({col(5 + si * 3)}{dv_row}*1,0)" for si in range(MAX_STAGES)
    )
    total_prop = "+".join(
        f"IFERROR({col(5+si*3)}{prop_mass_row},0)" for si in range(MAX_STAGES)
    )

    summary_rows = [
        ("Max Payload to LEO",          "kg",   f"=B{max_payload_row}"),
        ("Max Payload to GTO",          "kg",   f"=B{max_payload_gto_row}"),
        ("LEO Payload Fraction (GLOW)", "%",    f"=IFERROR(B{max_payload_row}/{col(5)}{m0_row}*100, \"\")"),
        ("Total Vehicle ΔV (actual)",   "m/s",  f"={dv_actual_sum}"),
        ("LEO ΔV Margin vs Mission",    "m/s",  f"={dv_actual_sum}-B{dv_budget_row}"),
        ("Gross Liftoff Mass (GLOW)",   "kg",   f"=IFERROR({col(5)}{m0_row}, \"\")"),
        ("Total Propellant Mass",       "kg",   f"={total_prop}"),
    ]

    for label, unit, formula in summary_rows:
        is_primary = label in ("Max Payload to LEO", "Max Payload to GTO")
        ws.cell(r, 1, value=label).font = Font(bold=True, size=11 if is_primary else 10)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=C["pos"] if is_primary else C["sec"])
        ws.cell(r, 2, value=formula).font = Font(bold=True, size=11 if is_primary else 10)
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=C["pos"] if is_primary else C["calc"])
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        ws.cell(r, 2).number_format = "#,##0" if unit in ("m/s", "kg") else "0.00"
        ws.cell(r, 3, value=unit).font = Font(size=9, color="777777")
        ws.cell(r, 3).alignment = Alignment(horizontal="left")
        apply_border_range(ws, r, r, 1, 3)
        r += 1

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Reference Comparison
# ─────────────────────────────────────────────────────────────────────────────
def build_comparison(wb):
    ws = wb.create_sheet("Comparison")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    style_hdr(ws, 1, 1, "REFERENCE VEHICLE QUICK-COMPARE — Key Performance Metrics", span=9)
    headers = ["Vehicle", "Stages", "GLOW (kg)", "Payload LEO (kg)",
               "Payload Fraction (%)", "Payload GTO (kg)", "Propulsion", "Height (m)", "Status"]
    widths   = [26, 8, 16, 16, 16, 16, 22, 10, 16]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_hdr(ws, 2, ci, h, dark=False)
        ws.column_dimensions[col(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    for ri, v in enumerate(VEHICLES, 3):
        pf = round(v["payload_leo_kg"] / v["glow_kg"] * 100, 2) if v["glow_kg"] and v["payload_leo_kg"] else None
        row_vals = [
            v["name"], len(v["stages"]), v["glow_kg"], v["payload_leo_kg"],
            pf, v.get("payload_gto_kg", "—"), v.get("propulsion", "—"), v["height_m"], v["status"]
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(ri, ci, value=val if val is not None else "—")
            cell.font = Font(bold=(ci == 1), size=10)
            cell.alignment = Alignment(horizontal="left" if ci in (1, 7, 9) else "center",
                                       vertical="center", indent=1)
            cell.fill = PatternFill("solid", fgColor=C["ltgray"] if ri % 2 == 0 else C["white"])
            if ci in (3, 4, 6):
                cell.number_format = "#,##0"
            if ci == 5 and val is not None:
                cell.number_format = "0.00"
        apply_border_range(ws, ri, ri, 1, len(headers))
        ws.row_dimensions[ri].height = 15

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Instructions
# ─────────────────────────────────────────────────────────────────────────────
def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    style_hdr(ws, 1, 1, "LAUNCH VEHICLE DESIGN TOOL — User Guide")
    ws.row_dimensions[1].height = 28

    lines = [
        ("HOW TO USE", True),
        ("1.  Go to 'Vehicle Design' tab — this is your main working sheet.", False),
        ("2.  Fill in Mission Requirements (yellow cells): payload mass, target orbit, delta-v budget.", False),
        ("3.  For each stage, enter the number of stages (1-3) and configure each stage:", False),
        ("      • Engine: Type the exact engine name from the 'Engine DB' tab. Isp will auto-populate via VLOOKUP.", False),
        ("      • Number of Engines: How many of that engine on this stage.", False),
        ("      • Propellant Combination: Must match exactly (e.g. LOX/RP-1, LOX/LH2, LOX/CH4, Solid HTPB, NTO/MMH).", False),
        ("      • Propellant Mass: Total usable propellant mass for this stage in kg.", False),
        ("      • Vac Isp Override: Leave blank to use the Engine DB value. Enter a number to override.", False),
        ("", False),
        ("4.  Subsystem masses auto-calculate from parametric models in the 'Propellants' tab.", False),
        ("      • Yellow cells in the subsystem section can be manually overridden with known values.", False),
        ("      • Use 'Dry Mass Override' if you have a known dry mass (e.g. from a reference vehicle).", False),
        ("", False),
        ("5.  Performance section calculates automatically:", False),
        ("      • Delta-V per stage via Tsiolkovsky equation: ΔV = Isp × g₀ × ln(m₀/m_f)", False),
        ("      • Burnout mass accounts for all upper stages + payload + payload adapter.", False),
        ("      • TWR uses sea-level thrust for Stage 1, vacuum thrust for upper stages.", False),
        ("", False),
        ("TIPS & RULES OF THUMB", True),
        ("• LEO missions typically need 9,000–9,500 m/s total ΔV (includes ~1,500 m/s gravity+drag losses).", False),
        ("• GTO missions need ~11,500–12,500 m/s depending on inclination.", False),
        ("• First stage TWR should be 1.2–1.5 at liftoff; too low = slow ascent; too high = unnecessary mass.", False),
        ("• Payload fraction for real vehicles: 1.5–4% to LEO is typical.", False),
        ("• Structural mass fraction (dry/propellant) is typically 5–15%; lower is better engineering.", False),
        ("• LOX/LH2 gives the highest Isp (~450s) but low density means large, heavy tanks.", False),
        ("• LOX/RP-1 gives good density and SL performance; workhorse of most first stages.", False),
        ("• LOX/CH4 is the emerging choice for reusability; good Isp + density balance.", False),
        ("", False),
        ("REFERENCE TABS", True),
        ("• Engine DB:   Real engine performance data for 35+ engines. Add your own rows at the bottom.", False),
        ("• Vehicle DB:  Real vehicle mass breakdowns for 11 vehicles. Use for design grounding.", False),
        ("• Propellants: Propellant properties + subsystem mass fraction model documentation.", False),
        ("• Comparison:  Quick performance comparison table across all reference vehicles.", False),
        ("", False),
        ("LIMITATIONS", True),
        ("• This tool uses simplified mass models. Real designs require detailed MDO and stress analysis.", False),
        ("• Delta-V calculations assume impulsive burns (no trajectory shaping or staging timing).", False),
        ("• Subsystem fractions are statistical — early concept accuracy is ±20–30%.", False),
        ("• Solid stages show N/A for SL Isp in engine DB — use Vac Isp as a conservative estimate.", False),
        ("• Parallel staging (e.g. strap-on boosters firing simultaneously with a core) is not modeled.", False),
        ("  Treat parallel-burn vehicles (Soyuz, Ariane 5) as series stages for delta-v calculation.", False),
    ]

    for ri, (text, bold) in enumerate(lines, 3):
        cell = ws.cell(ri, 1, value=text)
        cell.font = Font(bold=bold, size=10 if not bold else 11,
                         color="1F3864" if bold else "1F1F1F")
        if bold:
            cell.fill = PatternFill("solid", fgColor=C["sec"])
        ws.row_dimensions[ri].height = 16

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_readme(wb)
    build_design(wb)
    build_engine_db(wb)
    build_vehicle_db(wb)
    build_propellants(wb)
    build_comparison(wb)

    # Tab colors
    tab_colors = {
        "README":         "1F3864",
        "Vehicle Design": "2E7D32",
        "Engine DB":      "BF360C",
        "Vehicle DB":     "4527A0",
        "Propellants":    "00695C",
        "Comparison":     "1565C0",
    }
    for ws in wb.worksheets:
        if ws.title in tab_colors:
            ws.sheet_properties.tabColor = tab_colors[ws.title]

    wb.save(OUTPUT_FILE)
    print(f"✓ Saved: {OUTPUT_FILE}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"  Engines in DB: {len(ENGINES)}")
    print(f"  Reference vehicles: {len(VEHICLES)}")


if __name__ == "__main__":
    main()
